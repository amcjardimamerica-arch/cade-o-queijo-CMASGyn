#!/usr/bin/env python3
"""Coleta os repasses federais fundo a fundo do Fundo Nacional de Assistência Social.

Fonte: planilha oficial de extração de repasses, publicada em
fnas.mds.gov.br/extracao-dos-valores-repassados-fundo-a-fundo/

Traz, por município e competência: bloco, programa, valor e a ordem bancária do
SIAFI. É o lado da União na conciliação — o que foi efetivamente pago, contra o
que o Município registrou.

Vale registrar por que este módulo existe: os endereços antigos do Ministério
(aplicacoes.mds.gov.br) respondem 503 há meses. O portal novo publica o dado em
planilha, sem exigir autenticação. Quem procurar só pelos endereços antigos
conclui que o dado não existe.
"""
from __future__ import annotations
import json, os, sys, urllib.request
from collections import defaultdict
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
CACHE = RAIZ / "estado" / "fnas"
IBGE_GOIANIA = "520870"
UA = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) Chrome/120"}

PLANILHAS = {
 2020:"https://fnas.mds.gov.br/wp-content/uploads/2023/01/REPASSES_ANO_2020.xlsx",
 2021:"https://fnas.mds.gov.br/wp-content/uploads/2023/01/REPASSES_ANO_2021.xlsx",
 2022:"https://fnas.mds.gov.br/wp-content/uploads/2023/01/REPASSES_ANO_2022.xlsx",
 2023:"https://fnas.mds.gov.br/wp-content/uploads/2024/01/REPASSES_ANO_2023.xlsx",
 2024:"https://fnas.mds.gov.br/wp-content/uploads/2026/02/REPASSES_ANO_2024.xlsx",
 2025:"https://fnas.mds.gov.br/wp-content/uploads/2026/02/REPASSES_ANO_2025.xlsx",
}
# a planilha de um ano traz a competência do ano anterior na coluna ANO EXERCÍCIO;
# por isso o filtro é pela competência, não pelo nome do arquivo.

def baixar(ano):
    CACHE.mkdir(parents=True, exist_ok=True)
    dst = CACHE / f"repasses_{ano}.xlsx"
    if dst.exists() and dst.stat().st_size > 100000: return dst
    url = PLANILHAS.get(ano)
    if not url: return None
    try:
        req = urllib.request.Request(url, headers=UA)
        with urllib.request.urlopen(req, timeout=180) as r:
            dst.write_bytes(r.read())
        return dst
    except Exception as e:
        print(f"  ! {ano}: {e}", file=sys.stderr); return None

def ler(caminho, ibge=IBGE_GOIANIA):
    import openpyxl
    w = openpyxl.load_workbook(caminho, read_only=True)
    s = w[w.sheetnames[0]]
    out = []
    for i, r in enumerate(s.iter_rows(values_only=True)):
        if i == 0 or not r or str(r[0]) != ibge: continue
        if str(r[2]).upper() != "MUNICIPAL": continue
        try: v = float(r[9])
        except (TypeError, ValueError): continue
        out.append({"tipo": r[4], "bloco": r[5], "programa": r[6],
                    "competencia": f"{r[7]}-{str(r[8]).zfill(2)}",
                    "ano": str(r[7]), "mes": str(r[8]).zfill(2),
                    "valor": v, "ordem_bancaria": str(r[11])})
    w.close(); return out

def eh_igd(x):
    s = (str(x["bloco"]) + " " + str(x["programa"])).upper()
    return "IGD" in s

def main(anos=(2024, 2025)):
    tudo = []
    for a in anos:
        p = baixar(a)
        if not p: continue
        linhas = ler(p)
        tudo += linhas
        print(f"  {a}: {len(linhas)} repasses", file=sys.stderr)

    por_comp = defaultdict(float); por_bloco = defaultdict(float)
    igd_comp = defaultdict(float); igd_prog = defaultdict(float)
    for x in tudo:
        por_comp[x["competencia"]] += x["valor"]
        por_bloco[x["bloco"]] += x["valor"]
        if eh_igd(x):
            igd_comp[x["competencia"]] += x["valor"]
            igd_prog[x["programa"]] += x["valor"]

    # o piso de 10% incide sobre o repasse de CADA competência
    mensal = [{"competencia": k, "igd_repassado": round(v, 2),
               "devido_ao_controle_social": round(v * 0.10, 2),
               "piso_legal_3": round(v * 0.03, 2),
               "aplicado_publicado": None, "situacao": "SEM DEMONSTRATIVO PUBLICADO"}
              for k, v in sorted(igd_comp.items())]
    igd_total = sum(igd_comp.values())

    saida = {
     "fonte": {"nome": "Fundo Nacional de Assistência Social — extração de repasses fundo a fundo",
       "url": "https://fnas.mds.gov.br/extracao-dos-valores-repassados-fundo-a-fundo/",
       "municipio": "Goiânia", "codigo_ibge": IBGE_GOIANIA,
       "campos": "bloco, programa, competência, valor e ordem bancária do SIAFI"},
     "anos": list(anos), "repasses": len(tudo),
     "total_repassado": round(sum(x["valor"] for x in tudo), 2),
     "por_bloco": {k: round(v, 2) for k, v in sorted(por_bloco.items(), key=lambda i: -i[1])},
     "por_competencia": {k: round(v, 2) for k, v in sorted(por_comp.items())},
     "igd": {"total": round(igd_total, 2),
       "por_programa": {k: round(v, 2) for k, v in igd_prog.items()},
       "devido_ao_controle_social_10": round(igd_total * 0.10, 2),
       "competencias": mensal,
       "norma": "Artigo 6º da Resolução CNAS/MDS 202/2025",
       "nota": "O piso incide sobre o valor repassado em cada competência. Aplicar o "
               "percentual sobre o acumulado do ano não satisfaz a norma."},
     "detalhe": tudo}

    (RAIZ / "dados" / "repasses_federais.json").write_text(
        json.dumps(saida, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"  repasses federais: {len(tudo)} · total R$ {saida['total_repassado']:,.2f}")
    print(f"  IGD: R$ {igd_total:,.2f} · devido ao controle social R$ {igd_total*0.10:,.2f}")
    print(f"  competências com IGD: {len(mensal)} · com demonstrativo publicado: 0")

if __name__ == "__main__":
    main()
